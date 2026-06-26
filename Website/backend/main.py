"""
FastAPI backend — Telco Churn Prediction
Run: uvicorn main:app --reload --port 8000
"""
from fastapi import FastAPI, UploadFile, File, Query, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
import pandas as pd
import numpy as np
import joblib
import os
import io
import json
import tempfile
import uuid
import asyncio
from decimal import Decimal
from typing import Any, Optional, Set

from database import get_connection

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="ChurnIQ API", version="1.0.0")

# ── Dashboard cache (served instantly to every WS client) ─────────────────────
_dashboard_cache: Optional[dict] = None
_ws_clients: Set[WebSocket]      = set()

def _to_json(data: Any) -> str:
    """Serialize dict/list to JSON, converting Decimal → float so it never crashes."""
    def default(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
    return json.dumps(data, default=default)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
MODEL_DIR    = os.path.join(BASE_DIR, "..", "..", "ML", "models")
FRONTEND_DIR = os.path.join(BASE_DIR, "..", "frontend")

_model    = None
_scaler   = None
_features = None

def get_model():
    global _model, _scaler, _features
    if _model is None:
        _model    = joblib.load(os.path.join(MODEL_DIR, "Random_Forest_churn_model.pkl"))
        _scaler   = joblib.load(os.path.join(MODEL_DIR, "scaler.pkl"))
        _features = joblib.load(os.path.join(MODEL_DIR, "feature_names.pkl"))
    return _model, _scaler, _features


# ── Preprocessing (mirrors Phase 3 of the notebook) ──────────────────────────
def preprocess(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Total Charges"] = pd.to_numeric(df.get("Total Charges", 0), errors="coerce").fillna(0)

    df["Avg Monthly Spend"] = np.where(
        df["Tenure Months"] > 0,
        df["Total Charges"] / df["Tenure Months"],
        df["Monthly Charges"],
    )

    df["Tenure Band"] = pd.cut(
        df["Tenure Months"], bins=[0, 12, 24, 48, 72],
        labels=["0-12m", "13-24m", "25-48m", "49-72m"], right=True,
    )

    svc = ["Online Security", "Online Backup", "Device Protection",
           "Tech Support", "Streaming TV", "Streaming Movies"]
    df["Num Services"] = df[[c for c in svc if c in df.columns]].apply(
        lambda row: (row == "Yes").sum(), axis=1
    )

    bmap = {"Yes": 1, "No": 0, "Male": 1, "Female": 0}
    for col in ["Gender", "Senior Citizen", "Partner", "Dependents",
                "Phone Service", "Paperless Billing",
                "Online Security", "Online Backup", "Device Protection",
                "Tech Support", "Streaming TV", "Streaming Movies"]:
        if col in df.columns:
            df[col] = df[col].map(bmap).fillna(0).astype(int)

    nominal = [c for c in ["Multiple Lines", "Internet Service", "Contract", "Payment Method"]
               if c in df.columns]
    df = pd.get_dummies(df, columns=nominal, drop_first=False)

    df["Tenure Band"] = df["Tenure Band"].map(
        {"0-12m": 0, "13-24m": 1, "25-48m": 2, "49-72m": 3}
    ).fillna(0).astype(int)

    df = df.apply(pd.to_numeric, errors="coerce").fillna(0)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# Dashboard endpoints
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/stats/summary")
def get_summary():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    COUNT(*)                          AS total_customers,
                    SUM(churn_value)                  AS total_churned,
                    ROUND(AVG(monthly_charges), 2)    AS avg_monthly_charge,
                    ROUND(AVG(tenure_months), 1)      AS avg_tenure,
                    ROUND(SUM(total_charges), 2)      AS total_revenue,
                    ROUND(AVG(cltv), 0)               AS avg_cltv
                FROM customers
            """)
            row = cur.fetchone()
            row["churn_rate"] = round(row["total_churned"] / row["total_customers"] * 100, 1)
            return row
    finally:
        conn.close()


@app.get("/api/stats/churn-distribution")
def get_churn_dist():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT churn_label AS label, COUNT(*) AS value
                FROM customers GROUP BY churn_label
            """)
            return cur.fetchall()
    finally:
        conn.close()


@app.get("/api/stats/by-contract")
def by_contract():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    contract,
                    SUM(churn_value)                                          AS churned,
                    COUNT(*) - SUM(churn_value)                               AS retained,
                    ROUND(SUM(churn_value) / COUNT(*) * 100, 1)               AS churn_rate
                FROM customers
                GROUP BY contract
                ORDER BY churn_rate DESC
            """)
            return cur.fetchall()
    finally:
        conn.close()


@app.get("/api/stats/by-internet")
def by_internet():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    internet_service,
                    SUM(churn_value)                                          AS churned,
                    COUNT(*) - SUM(churn_value)                               AS retained,
                    ROUND(SUM(churn_value) / COUNT(*) * 100, 1)               AS churn_rate
                FROM customers
                GROUP BY internet_service
                ORDER BY churn_rate DESC
            """)
            return cur.fetchall()
    finally:
        conn.close()


@app.get("/api/stats/by-payment")
def by_payment():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    payment_method,
                    SUM(churn_value)                                          AS churned,
                    COUNT(*) - SUM(churn_value)                               AS retained,
                    ROUND(SUM(churn_value) / COUNT(*) * 100, 1)               AS churn_rate
                FROM customers
                GROUP BY payment_method
                ORDER BY churn_rate DESC
            """)
            return cur.fetchall()
    finally:
        conn.close()


@app.get("/api/stats/by-senior")
def by_senior():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    senior_citizen,
                    SUM(churn_value)                                          AS churned,
                    COUNT(*) - SUM(churn_value)                               AS retained,
                    ROUND(SUM(churn_value) / COUNT(*) * 100, 1)               AS churn_rate
                FROM customers
                GROUP BY senior_citizen
            """)
            return cur.fetchall()
    finally:
        conn.close()


@app.get("/api/stats/tenure-histogram")
def tenure_histogram():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    CASE
                        WHEN tenure_months <= 12 THEN '0-12'
                        WHEN tenure_months <= 24 THEN '13-24'
                        WHEN tenure_months <= 36 THEN '25-36'
                        WHEN tenure_months <= 48 THEN '37-48'
                        WHEN tenure_months <= 60 THEN '49-60'
                        ELSE '61-72'
                    END AS bin,
                    churn_label,
                    COUNT(*) AS count
                FROM customers
                GROUP BY bin, churn_label
                ORDER BY FIELD(bin,'0-12','13-24','25-36','37-48','49-60','61-72')
            """)
            rows = cur.fetchall()
            bins   = ['0-12','13-24','25-36','37-48','49-60','61-72']
            data   = {b: {'No': 0, 'Yes': 0} for b in bins}
            for r in rows:
                data[r['bin']][r['churn_label']] = r['count']
            return {
                "categories": bins,
                "retained":   [data[b]['No']  for b in bins],
                "churned":    [data[b]['Yes'] for b in bins],
            }
    finally:
        conn.close()


@app.get("/api/stats/monthly-charges-histogram")
def monthly_charges_histogram():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    CASE
                        WHEN monthly_charges < 30 THEN '$18-30'
                        WHEN monthly_charges < 45 THEN '$30-45'
                        WHEN monthly_charges < 60 THEN '$45-60'
                        WHEN monthly_charges < 75 THEN '$60-75'
                        WHEN monthly_charges < 90 THEN '$75-90'
                        ELSE '$90+'
                    END AS bin,
                    churn_label,
                    COUNT(*) AS count
                FROM customers
                GROUP BY bin, churn_label
                ORDER BY FIELD(bin,'$18-30','$30-45','$45-60','$60-75','$75-90','$90+')
            """)
            rows = cur.fetchall()
            bins = ['$18-30','$30-45','$45-60','$60-75','$75-90','$90+']
            data = {b: {'No': 0, 'Yes': 0} for b in bins}
            for r in rows:
                data[r['bin']][r['churn_label']] = r['count']
            return {
                "categories": bins,
                "retained":   [data[b]['No']  for b in bins],
                "churned":    [data[b]['Yes'] for b in bins],
            }
    finally:
        conn.close()


@app.get("/api/stats/churn-reasons")
def churn_reasons():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT churn_reason AS reason, COUNT(*) AS count
                FROM customers
                WHERE churn_reason IS NOT NULL AND churn_reason != ''
                GROUP BY churn_reason
                ORDER BY count DESC
                LIMIT 10
            """)
            return cur.fetchall()
    finally:
        conn.close()


@app.get("/api/stats/customers")
def get_customers(
    page:    int = Query(1, ge=1),
    per_page:int = Query(10, ge=5, le=100),
    search:  str = Query(""),
    churn:   str = Query(""),
):
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            where_parts = []
            params      = []

            if search:
                where_parts.append(
                    "(customer_id LIKE %s OR contract LIKE %s OR internet_service LIKE %s)"
                )
                params += [f"%{search}%"] * 3

            if churn in ("Yes", "No"):
                where_parts.append("churn_label = %s")
                params.append(churn)

            where_sql = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

            cur.execute(f"SELECT COUNT(*) AS n FROM customers {where_sql}", params)
            total = cur.fetchone()["n"]

            offset = (page - 1) * per_page
            cur.execute(f"""
                SELECT customer_id, gender, senior_citizen, tenure_months,
                       contract, internet_service, monthly_charges,
                       total_charges, churn_label, churn_score
                FROM customers {where_sql}
                ORDER BY id
                LIMIT %s OFFSET %s
            """, params + [per_page, offset])

            return {
                "data":     cur.fetchall(),
                "total":    total,
                "page":     page,
                "per_page": per_page,
                "pages":    (total + per_page - 1) // per_page,
            }
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# WebSocket — real-time dashboard feed
# ═══════════════════════════════════════════════════════════════════════════════

def collect_dashboard_data() -> dict:
    """Run every dashboard query in one DB connection and return a single payload."""
    conn = get_connection()
    try:
        with conn.cursor() as cur:

            # Summary KPIs
            cur.execute("""
                SELECT COUNT(*) AS total_customers, SUM(churn_value) AS total_churned,
                    ROUND(AVG(monthly_charges),2) AS avg_monthly_charge,
                    ROUND(AVG(tenure_months),1)   AS avg_tenure,
                    ROUND(SUM(total_charges),2)   AS total_revenue,
                    ROUND(AVG(cltv),0)             AS avg_cltv
                FROM customers""")
            summary = cur.fetchone()
            summary["churn_rate"] = round(
                summary["total_churned"] / summary["total_customers"] * 100, 1)

            # Churn distribution
            cur.execute("""SELECT churn_label AS label, COUNT(*) AS value
                FROM customers GROUP BY churn_label""")
            churn_dist = cur.fetchall()

            # By contract
            cur.execute("""SELECT contract,
                SUM(churn_value) AS churned, COUNT(*)-SUM(churn_value) AS retained,
                ROUND(SUM(churn_value)/COUNT(*)*100,1) AS churn_rate
                FROM customers GROUP BY contract ORDER BY churn_rate DESC""")
            by_contract = cur.fetchall()

            # By internet
            cur.execute("""SELECT internet_service,
                SUM(churn_value) AS churned, COUNT(*)-SUM(churn_value) AS retained,
                ROUND(SUM(churn_value)/COUNT(*)*100,1) AS churn_rate
                FROM customers GROUP BY internet_service ORDER BY churn_rate DESC""")
            by_internet = cur.fetchall()

            # By payment
            cur.execute("""SELECT payment_method,
                SUM(churn_value) AS churned, COUNT(*)-SUM(churn_value) AS retained,
                ROUND(SUM(churn_value)/COUNT(*)*100,1) AS churn_rate
                FROM customers GROUP BY payment_method ORDER BY churn_rate DESC""")
            by_payment = cur.fetchall()

            # By senior citizen
            cur.execute("""SELECT senior_citizen,
                SUM(churn_value) AS churned, COUNT(*)-SUM(churn_value) AS retained,
                ROUND(SUM(churn_value)/COUNT(*)*100,1) AS churn_rate
                FROM customers GROUP BY senior_citizen""")
            by_senior = cur.fetchall()

            # Tenure histogram
            cur.execute("""
                SELECT CASE
                    WHEN tenure_months <= 12 THEN '0-12'
                    WHEN tenure_months <= 24 THEN '13-24'
                    WHEN tenure_months <= 36 THEN '25-36'
                    WHEN tenure_months <= 48 THEN '37-48'
                    WHEN tenure_months <= 60 THEN '49-60'
                    ELSE '61-72' END AS bin,
                    churn_label, COUNT(*) AS count
                FROM customers GROUP BY bin, churn_label
                ORDER BY FIELD(bin,'0-12','13-24','25-36','37-48','49-60','61-72')""")
            t_rows = cur.fetchall()
            t_bins = ['0-12','13-24','25-36','37-48','49-60','61-72']
            t_data = {b: {'No': 0, 'Yes': 0} for b in t_bins}
            for r in t_rows:
                t_data[r['bin']][r['churn_label']] = r['count']
            tenure_hist = {
                'categories': t_bins,
                'retained':   [t_data[b]['No']  for b in t_bins],
                'churned':    [t_data[b]['Yes'] for b in t_bins],
            }

            # Monthly charges histogram
            cur.execute("""
                SELECT CASE
                    WHEN monthly_charges < 30 THEN '$18-30'
                    WHEN monthly_charges < 45 THEN '$30-45'
                    WHEN monthly_charges < 60 THEN '$45-60'
                    WHEN monthly_charges < 75 THEN '$60-75'
                    WHEN monthly_charges < 90 THEN '$75-90'
                    ELSE '$90+' END AS bin,
                    churn_label, COUNT(*) AS count
                FROM customers GROUP BY bin, churn_label
                ORDER BY FIELD(bin,'$18-30','$30-45','$45-60','$60-75','$75-90','$90+')""")
            c_rows = cur.fetchall()
            c_bins = ['$18-30','$30-45','$45-60','$60-75','$75-90','$90+']
            c_data = {b: {'No': 0, 'Yes': 0} for b in c_bins}
            for r in c_rows:
                c_data[r['bin']][r['churn_label']] = r['count']
            charges_hist = {
                'categories': c_bins,
                'retained':   [c_data[b]['No']  for b in c_bins],
                'churned':    [c_data[b]['Yes'] for b in c_bins],
            }

            # Churn reasons (top 10)
            cur.execute("""SELECT churn_reason AS reason, COUNT(*) AS count
                FROM customers
                WHERE churn_reason IS NOT NULL AND churn_reason != ''
                GROUP BY churn_reason ORDER BY count DESC LIMIT 10""")
            reasons = cur.fetchall()

            return {
                'summary':          summary,
                'churn_distribution': churn_dist,
                'by_contract':      by_contract,
                'by_internet':      by_internet,
                'by_payment':       by_payment,
                'by_senior':        by_senior,
                'tenure_histogram': tenure_hist,
                'charges_histogram':charges_hist,
                'churn_reasons':    reasons,
            }
    finally:
        conn.close()


async def _push_all(data: dict) -> None:
    """Broadcast a payload to every connected WS client; remove dead ones."""
    dead = set()
    payload = _to_json(data)
    for ws in list(_ws_clients):
        try:
            await ws.send_text(payload)
        except Exception:
            dead.add(ws)
    _ws_clients.difference_update(dead)


async def _cache_refresh_loop() -> None:
    """Background task: refresh cache every 30 s and push to all clients."""
    global _dashboard_cache
    while True:
        await asyncio.sleep(30)
        try:
            new_data = await asyncio.to_thread(collect_dashboard_data)
            _dashboard_cache = new_data
            await _push_all(new_data)
        except Exception:
            pass


@app.on_event("startup")
async def _startup() -> None:
    """Pre-warm the dashboard cache so the first WS hit is instant."""
    global _dashboard_cache
    try:
        _dashboard_cache = await asyncio.to_thread(collect_dashboard_data)
    except Exception:
        pass
    asyncio.create_task(_cache_refresh_loop())


@app.websocket("/ws/dashboard")
async def dashboard_ws(websocket: WebSocket):
    await websocket.accept()
    _ws_clients.add(websocket)
    try:
        # Use cached data (instant) or fall back to a fresh DB query
        data = _dashboard_cache
        if data is None:
            data = await asyncio.to_thread(collect_dashboard_data)
        await websocket.send_text(_to_json(data))
        # Hold the connection open; refreshes are pushed by _cache_refresh_loop
        await websocket.receive_text()
    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        _ws_clients.discard(websocket)


# ═══════════════════════════════════════════════════════════════════════════════
# Prediction endpoints
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/predict/template")
def download_template():
    """Return a blank Excel template with the required input columns."""
    cols = [
        "Gender", "Senior Citizen", "Partner", "Dependents",
        "Tenure Months", "Phone Service", "Multiple Lines",
        "Internet Service", "Online Security", "Online Backup",
        "Device Protection", "Tech Support", "Streaming TV",
        "Streaming Movies", "Contract", "Paperless Billing",
        "Payment Method", "Monthly Charges", "Total Charges",
    ]
    sample = {
        "Gender": "Female", "Senior Citizen": "No", "Partner": "No",
        "Dependents": "No", "Tenure Months": 2, "Phone Service": "Yes",
        "Multiple Lines": "No", "Internet Service": "Fiber optic",
        "Online Security": "No", "Online Backup": "No",
        "Device Protection": "No", "Tech Support": "No",
        "Streaming TV": "No", "Streaming Movies": "No",
        "Contract": "Month-to-month", "Paperless Billing": "Yes",
        "Payment Method": "Electronic check",
        "Monthly Charges": 70.70, "Total Charges": 151.65,
    }
    df = pd.DataFrame([sample], columns=cols)
    path = os.path.join(tempfile.gettempdir(), "churn_template.xlsx")
    df.to_excel(path, index=False)
    return FileResponse(path, filename="churn_prediction_template.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/predict/upload")
async def predict_upload(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "Only .xlsx / .xls / .csv files are supported.")

    content = await file.read()

    try:
        if file.filename.endswith(".csv"):
            df_input = pd.read_csv(io.BytesIO(content))
        else:
            df_input = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    if df_input.empty:
        raise HTTPException(400, "Uploaded file has no rows.")

    # Strip extra/result columns if user re-uploads a result file
    drop_if_present = ["Churn Prediction", "Churn Probability %", "Risk Level", "Recommendation",
                       "Churn Label", "Churn Value", "Churn Score", "CLTV", "Churn Reason",
                       "CustomerID", "Count", "Country", "State", "City",
                       "Zip Code", "Lat Long", "Latitude", "Longitude"]
    df_work = df_input.drop(columns=[c for c in drop_if_present if c in df_input.columns])

    required = ["Tenure Months", "Monthly Charges"]
    missing  = [c for c in required if c not in df_work.columns]
    if missing:
        raise HTTPException(400, f"Required columns missing: {missing}")

    model, scaler, feature_names = get_model()

    df_proc = preprocess(df_work)

    # Align columns to training features
    for col in feature_names:
        if col not in df_proc.columns:
            df_proc[col] = 0
    df_proc = df_proc[feature_names].apply(pd.to_numeric, errors="coerce").fillna(0)

    probs = model.predict_proba(df_proc)[:, 1]
    preds = (probs >= 0.5).astype(int)

    def risk(p):
        if p >= 0.70: return "HIGH"
        if p >= 0.40: return "MEDIUM"
        return "LOW"

    def action(p):
        if p >= 0.70: return "DROP / Intervene Now"
        if p >= 0.40: return "MONITOR"
        return "KEEP"

    df_out = df_input.copy()
    df_out["Churn Prediction"]  = preds
    df_out["Churn Probability %"] = (probs * 100).round(1)
    df_out["Risk Level"]          = [risk(p)   for p in probs]
    df_out["Recommendation"]      = [action(p) for p in probs]

    # Save result Excel
    result_path = os.path.join(
        tempfile.gettempdir(), f"churn_results_{uuid.uuid4().hex[:8]}.xlsx"
    )
    with pd.ExcelWriter(result_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Predictions")
        wb   = writer.book
        ws   = wb["Predictions"]
        from openpyxl.styles import PatternFill, Font
        red_fill   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        pred_col = df_out.columns.get_loc("Churn Prediction") + 1
        for row_idx, pred in enumerate(preds, start=2):
            fill = red_fill if pred == 1 else green_fill
            for col_idx in range(1, len(df_out.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Build JSON preview (first 200 rows for UI table)
    preview = df_out.head(200).replace({np.nan: None}).to_dict("records")

    summary = {
        "total":         len(preds),
        "churn_count":   int(preds.sum()),
        "keep_count":    int((preds == 0).sum()),
        "high_risk":     int(sum(p >= 0.70 for p in probs)),
        "medium_risk":   int(sum(0.40 <= p < 0.70 for p in probs)),
        "low_risk":      int(sum(p < 0.40 for p in probs)),
        "churn_rate":    round(float(probs.mean()) * 100, 1),
        "download_path": result_path,
    }

    return {"summary": summary, "preview": preview}


@app.get("/api/predict/download")
def download_result(path: str = Query(...)):
    if not os.path.exists(path):
        raise HTTPException(404, "Result file not found. Please re-run prediction.")
    return FileResponse(
        path, filename="churn_prediction_results.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Serve frontend (must be last — API routes take priority) ──────────────────
app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
